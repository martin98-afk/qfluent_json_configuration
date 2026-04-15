# -*- coding: utf-8 -*-
"""
工况转换脚本 - 根据专家规则生成寻优上下限

专家规则:
    1. 以"摆嘴"为判断条件，摆嘴数值相加≥1表示该铁口打开
    2. 统计三个铁口打开的数量
    3. 根据打开数量设置基础值:
       - 0个铁口打开: 750
       - 1个铁口打开: 800
       - 2个铁口打开: 840
       - 3个铁口打开: 880
    4. 非当前打开铁口的侧吸和顶吸数值，每增加1个，数值多增加10
"""

import openpyxl
import os
import sys
import argparse
from typing import List, Tuple, Optional

# ============================================================
# 标签映射表 - 原始测点名 → 简化标签
# ============================================================
LABEL_MAPPING = {
    "2#高炉出铁场2#除尘风机 | AN_LK_2GL_FJ_7_01 | 1#铁口侧吸除尘阀门（CC1_V03)开信号": "1#铁口侧吸",
    "2#高炉出铁场2#除尘风机 | AN_LK_2GL_FJ_7_03 | 1#铁口出铁口平台下除尘阀门（CC1_V05)开信号": "1#铁口摆嘴",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_04 | 1#铁口摆动流嘴除尘阀门（CC1_V11)开信号": "1#铁口摆嘴",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_06 | 1#铁口顶吸除尘阀门（CC1_V02)开信号": "1#铁口顶吸",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_07 | 1#铁口顶吸除尘阀门（CC1_V01)开信号": "1#铁口顶吸",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_13 | 2#铁口侧吸除尘阀门（CC2_V04)开信号": "2#铁口侧吸",
    "2#高炉出铁场2#除尘风机 | AN_LK_2GL_FJ_7_14 | 2#铁口出铁口平台下除尘阀门（CC2_V05)开信号": "2#铁口摆嘴",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_15 | 2#铁口顶吸除尘阀门（CC2_V02)开信号": "2#铁口顶吸",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_16 | 2#铁口顶吸除尘阀门（CC2_V01)开信号": "2#铁口顶吸",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_19 | 3#铁口侧吸除尘阀门（CC3_V03)开信号": "3#铁口侧吸",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_20 | 3#铁口侧吸除尘阀门（CC3_V04)开信号": "3#铁口侧吸",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_21 | 3#铁口摆动流嘴除尘阀门（CC3_V11)开信号": "3#铁口摆嘴",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_22 | 3#铁口出铁口平台下除尘阀门（CC3_V05)开信号": "3#铁口摆嘴",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_24 | 3#铁口顶吸除尘阀门（CC3_V02)开信号": "3#铁口顶吸",
    "2#高炉出铁场1#除尘风机 | AN_LK_2GL_FJ_7_25 | 3#铁口顶吸除尘阀门（CC3_V01)开信号": "3#铁口顶吸",
}

# ============================================================
# 列索引定义 (0-based，从第2列开始，对应Excel列B,C,...)
# ============================================================

# 摆嘴列索引 - 用于判断铁口是否打开
PORT_NOZZLE_COLS = {
    "1#铁口": [3, 4],  # 1#铁口摆嘴列
    "2#铁口": [8],      # 2#铁口摆嘴列
    "3#铁口": [14, 15]  # 3#铁口摆嘴列
}

# 侧吸列索引
PORT_SIDESUCTION_COLS = {
    "1#铁口": [1, 2],  # 1#铁口侧吸列
    "2#铁口": [7],      # 2#铁口侧吸列
    "3#铁口": [11, 12]  # 3#铁口侧吸列
}

# 顶吸列索引
PORT_TOPSUCTION_COLS = {
    "1#铁口": [5, 6],  # 1#铁口顶吸列
    "2#铁口": [9, 10],  # 2#铁口顶吸列
    "3#铁口": [16, 17]  # 3#铁口顶吸列
}

# ============================================================
# 基础值定义 - 根据打开铁口数量
# ============================================================
BASE_VALUES = {
    0: 750,  # 0个铁口打开
    1: 800,  # 1个铁口打开
    2: 840,  # 2个铁口打开
    3: 880   # 3个铁口打开
}


def safe_to_number(value) -> float:
    """
    安全地将值转换为数字类型
    
    Args:
        value: 任意类型的值
    
    Returns:
        数字值，如果是None或无法转换则返回0
    """
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if value == '':
            return 0
        try:
            return float(value)
        except ValueError:
            return 0
    return 0


def calculate_limits(row_data: List) -> Tuple[float, float]:
    """
    根据专家规则计算寻优上限和下限
    
    规则说明:
        1. 以"摆嘴"为判断条件，摆嘴数值相加≥1表示该铁口打开
        2. 统计三个铁口打开的数量
        3. 根据打开数量设置基础值
        4. 非当前打开铁口的侧吸和顶吸数值，每增加1个，数值多增加10
    
    Args:
        row_data: 行数据列表
    
    Returns:
        (上限, 下限) 元组
    """
    open_ports = []
    
    # 步骤1: 判断每个铁口是否打开 (摆嘴数值相加≥1)
    for port, nozzle_cols in PORT_NOZZLE_COLS.items():
        nozzle_sum = 0
        for col in nozzle_cols:
            if col < len(row_data) and row_data[col] is not None:
                nozzle_sum += safe_to_number(row_data[col])
        if nozzle_sum >= 1:
            open_ports.append(port)
    
    # 步骤2: 统计打开的铁口数量，设置基础值
    num_open = len(open_ports)
    base_value = BASE_VALUES.get(num_open, 750)
    
    # 步骤3: 计算非当前打开铁口的侧吸和顶吸额外值
    extra_value = 0
    all_ports = ["1#铁口", "2#铁口", "3#铁口"]
    
    for port in all_ports:
        if port not in open_ports:
            # 计算侧吸数值
            side_sum = 0
            for col in PORT_SIDESUCTION_COLS[port]:
                if col < len(row_data) and row_data[col] is not None:
                    side_sum += safe_to_number(row_data[col])
            
            # 计算顶吸数值
            top_sum = 0
            for col in PORT_TOPSUCTION_COLS[port]:
                if col < len(row_data) and row_data[col] is not None:
                    top_sum += safe_to_number(row_data[col])
            
            # 每增加1个侧吸或顶吸，数值多增加10
            if side_sum >= 1:
                extra_value += 10
            if top_sum >= 1:
                extra_value += 10
    
    final_value = base_value + extra_value
    return final_value, final_value


def process_excel_file(input_filepath: str, output_filepath: Optional[str] = None) -> None:
    """
    处理Excel文件，根据专家规则更新寻优上限和下限
    
    Args:
        input_filepath: 输入文件路径
        output_filepath: 输出文件路径，默认自动生成新文件名（原文件后添加_processed）
    """
    if not output_filepath:
        # 默认输出到新文件，不覆盖原文件
        name, ext = os.path.splitext(input_filepath)
        output_filepath = f"{name}_processed{ext}"
    
    wb = openpyxl.load_workbook(input_filepath)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        
        # 查找"寻优上限"和"寻优下限"列索引
        upper_col_idx = None
        lower_col_idx = None
        
        for col_idx, header in enumerate(headers):
            if header == "寻优上限":
                upper_col_idx = col_idx + 1  # 转为1-based
            elif header == "寻优下限":
                lower_col_idx = col_idx + 1  # 转为1-based
        
        if upper_col_idx is None or lower_col_idx is None:
            print(f"警告: Sheet '{sheet_name}' 中未找到'寻优上限'或'寻优下限'列，跳过")
            continue
        
        # 处理每一行数据
        for row_idx in range(2, ws.max_row + 1):
            row_data = [ws.cell(row=row_idx, column=col + 1).value for col in range(len(headers))]
            upper, lower = calculate_limits(row_data)
            
            ws.cell(row=row_idx, column=upper_col_idx).value = upper
            ws.cell(row=row_idx, column=lower_col_idx).value = lower
    
    wb.save(output_filepath)
    print(f"处理完成: {input_filepath} -> {output_filepath}")


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description="工况转换脚本 - 根据专家规则生成寻优上下限")
    parser.add_argument("input_file", nargs="?", default=None, help="输入Excel文件路径")
    parser.add_argument("output_file", nargs="?", default=None, help="输出Excel文件路径(可选，默认生成新文件)")
    args = parser.parse_args()
    
    if args.input_file:
        # 处理指定文件，输出到新文件
        process_excel_file(args.input_file, args.output_file)
    else:
        # 扫描当前目录下所有 output*.xlsx 文件
        script_dir = os.path.dirname(os.path.abspath(__file__))
        processed_count = 0
        
        for filename in os.listdir(script_dir):
            if filename.startswith("output") and filename.endswith(".xlsx"):
                filepath = os.path.join(script_dir, filename)
                process_excel_file(filepath)
                processed_count += 1
        
        if processed_count == 0:
            print("未找到需要处理的文件 (output*.xlsx)")
        else:
            print(f"共处理 {processed_count} 个文件")


if __name__ == "__main__":
    main()
